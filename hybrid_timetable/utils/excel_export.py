from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.cell import MergedCell
from .clashes import highlight_clashes

def export_schedule_to_excel(schedule, days, slots_per_day, filename="timetable.xlsx",
                             time_labels=None, semester_name="Semester II",
                             clashes=None, groups=None, faculties=None,
                             room_matrix=None, section="A", dept="CSE-AI & ML",
                             room_no="8202", ay="2025-26", section_incharge=None):
    wb = Workbook()

    # Default time labels (your real slots)
    if time_labels is None:
        time_labels = [
            "09:00-09:50","09:50-10:40","10:40-11:30","11:30-12:20",
            "12:20-01:10","01:10-01:55","01:55-02:40","02:40-03:25","03:25-04:10"
        ][:slots_per_day]

    def make_sheet(ws, schedule_subset, title, subtitle=None):
        ws.title = title

        # Top header (college info)
        # Removed institution details as per user request
        # ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(days)+1)
        # ws.cell(row=1, column=1, value="SREENIDHI INSTITUTE OF SCIENCE AND TECHNOLOGY").font = Font(bold=True, size=14)
        # ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        # ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=len(days)+1)
        # ws.cell(row=2, column=1, value=f"DEPARTMENT OF {dept} | A.Y. {ay} | Section {section} | Room No: {room_no}").font = Font(bold=True, size=12)
        # ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

        # Table header
        ws.cell(row=3, column=1, value="Slot/Day").font = Font(bold=True)
        for j, day in enumerate(days, start=2):
            ws.cell(row=3, column=j, value=day).font = Font(bold=True)

        # Slot labels + lunch row
        row_offset = 4
        for slot, label in enumerate(time_labels, start=1):
            ws.cell(row=row_offset + slot - 1, column=1, value=f"Slot {slot}\n{label}")

        # Fill schedule
        for sess_id, info in sorted(schedule_subset.items(), key=lambda x: x[1]["start"]):
            start = info["start"]
            length = info["length"]
            room = info["room"]
            meta = info["meta"]

            day_idx = start // slots_per_day
            slot_idx = start % slots_per_day

            text = f"{meta['name']} ({meta['faculty']})\n{room}\n{meta['group']}"

            r1 = row_offset + slot_idx
            r2 = r1 + length - 1
            c = day_idx + 2

            # Get the target cell
            cell = ws.cell(row=r1, column=c)

            # Check if it's already in a merged range
            for merged_range in list(ws.merged_cells.ranges):
                if cell.coordinate in merged_range:
                    # Unmerge so we can write safely
                    ws.unmerge_cells(str(merged_range))
                    cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    break

            # Now safe to write
            cell.value = text
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

            # Merge multi-slot labs after writing
            if length > 1:
                ws.merge_cells(start_row=r1, end_row=r2, start_column=c, end_column=c)

            cname = meta['name'].lower()
            if "lab" in cname:
                cell.fill = PatternFill(start_color="FFD9E6", end_color="FFD9E6", fill_type="solid")
            elif "sports" in cname or "library" in cname or "mentoring" in cname:
                cell.fill = PatternFill(start_color="D9FFD9", end_color="D9FFD9", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

        # Footer (Section Incharge)
        if section_incharge:
            footer_row = row_offset + slots_per_day + 2
            merge_range = f"A{footer_row}:{chr(65 + len(days))} {footer_row}"
            if not any(merge_range == str(mr) for mr in ws.merged_cells.ranges):
                ws.merge_cells(start_row=footer_row, end_row=footer_row,
                               start_column=1, end_column=len(days)+1)
            ws.cell(row=footer_row, column=1,
                    value=f"Section Incharge: {section_incharge}").font = Font(bold=True)

        # Adjust column width
        for col in ws.columns:
            max_len = 0
            col_letter = None
            for cell in col:
                if isinstance(cell, MergedCell): continue
                col_letter = cell.column_letter
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            if col_letter:
                ws.column_dimensions[col_letter].width = max_len + 4

    # Create Index Sheet
    ws_index = wb.active
    ws_index.title = "Index"
    ws_index.cell(row=1, column=1, value="TIMETABLE INDEX - SREENIDHI INSTITUTE").font = Font(bold=True, size=16)
    ws_index.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    # Check if cells are already merged before merging
    merge_range = f"A1:C1"
    if not any(merge_range == str(mr) for mr in ws_index.merged_cells.ranges):
        ws_index.merge_cells(start_row=1, end_row=1, start_column=1, end_column=3)

    ws_index.cell(row=3, column=1, value="SECTION").font = Font(bold=True)
    ws_index.cell(row=3, column=2, value="SHEET NAME").font = Font(bold=True)
    ws_index.cell(row=3, column=3, value="DESCRIPTION").font = Font(bold=True)

    row_idx = 4

    # Master Timetable Section
    ws_index.cell(row=row_idx, column=1, value="📊 MASTER TIMETABLE").font = Font(bold=True, color="FF0000")
    ws_index.cell(row=row_idx, column=2, value="Master")
    ws_index.cell(row=row_idx, column=3, value="Complete timetable overview")
    row_idx += 1

    # Group-Based Timetables Section
    ws_index.cell(row=row_idx, column=1, value="👥 GROUP-BASED TIMETABLES").font = Font(bold=True, color="0000FF")
    row_idx += 1

    groups_in_schedule = {info["meta"]["group"] for info in schedule.values()}
    for g in sorted(groups_in_schedule):
        ws_index.cell(row=row_idx, column=2, value=f"Group_{g}")
        ws_index.cell(row=row_idx, column=3, value=f"Timetable for Group {g}")
        row_idx += 1

    # Faculty-Based Timetables Section
    ws_index.cell(row=row_idx, column=1, value="👨‍🏫 FACULTY-BASED TIMETABLES").font = Font(bold=True, color="008000")
    row_idx += 1

    faculties_in_schedule = {info["meta"]["faculty"] for info in schedule.values()}
    for f in sorted(faculties_in_schedule):
        ws_index.cell(row=row_idx, column=2, value=f"Faculty_{f}")
        ws_index.cell(row=row_idx, column=3, value=f"Timetable for {f}")
        row_idx += 1

    # Room-Based Timetables Section
    ws_index.cell(row=row_idx, column=1, value="🏫 ROOM-BASED TIMETABLES").font = Font(bold=True, color="800080")
    row_idx += 1

    rooms_in_schedule = {info["room"] for info in schedule.values()}
    for r in sorted(rooms_in_schedule):
        ws_index.cell(row=row_idx, column=2, value=f"Room_{r}")
        ws_index.cell(row=row_idx, column=3, value=f"Room {r} utilization")
        row_idx += 1

    # Analysis Section
    ws_index.cell(row=row_idx, column=1, value="📈 ANALYSIS & STATISTICS").font = Font(bold=True, color="FFA500")
    row_idx += 1
    ws_index.cell(row=row_idx, column=2, value="Statistics")
    ws_index.cell(row=row_idx, column=3, value="Timetable statistics and metrics")
    row_idx += 1
    ws_index.cell(row=row_idx, column=2, value="Clash_Analysis")
    ws_index.cell(row=row_idx, column=3, value="Conflict analysis report")

    # Adjust column widths for index
    ws_index.column_dimensions['A'].width = 25
    ws_index.column_dimensions['B'].width = 20
    ws_index.column_dimensions['C'].width = 40

    # ===== MASTER TIMETABLE =====
    ws_master = wb.create_sheet(title="Master")
    make_sheet(ws_master, schedule, "Master", subtitle=f"{semester_name} - Complete Timetable")

    if room_matrix is not None:
        highlight_clashes(ws_master, room_matrix)

    # ===== GROUP-BASED TIMETABLES =====
    groups_in_schedule = {info["meta"]["group"] for info in schedule.values()}
    for g in sorted(groups_in_schedule):
        ws = wb.create_sheet(title=f"Group_{g}")
        subset = {sid: info for sid, info in schedule.items() if info["meta"]["group"] == g}
        make_sheet(ws, subset, f"Group {g}", subtitle=f"{semester_name} - Group {g} Timetable")

    # ===== FACULTY-BASED TIMETABLES =====
    faculties_in_schedule = {info["meta"]["faculty"] for info in schedule.values()}
    for f in sorted(faculties_in_schedule):
        ws = wb.create_sheet(title=f"Faculty_{f}")
        subset = {sid: info for sid, info in schedule.items() if info["meta"]["faculty"] == f}
        make_sheet(ws, subset, f"Faculty {f}", subtitle=f"{semester_name} - {f} Schedule")

    # ===== ROOM-BASED TIMETABLES =====
    rooms_in_schedule = {info["room"] for info in schedule.values()}
    for r in sorted(rooms_in_schedule):
        ws = wb.create_sheet(title=f"Room_{r}")
        subset = {sid: info for sid, info in schedule.items() if info["room"] == r}
        make_sheet(ws, subset, f"Room {r}", subtitle=f"{semester_name} - Room {r} Utilization")

    # ===== STATISTICS SHEET =====
    ws_stats = wb.create_sheet(title="Statistics")
    ws_stats.cell(row=1, column=1, value="TIMETABLE STATISTICS").font = Font(bold=True, size=16)
    ws_stats.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    # Check if cells are already merged before merging
    merge_range = "A1:D1"
    if not any(merge_range == str(mr) for mr in ws_stats.merged_cells.ranges):
        ws_stats.merge_cells(start_row=1, end_row=1, start_column=1, end_column=4)

    # Basic Statistics
    ws_stats.cell(row=3, column=1, value="📊 BASIC STATISTICS").font = Font(bold=True, size=14)
    ws_stats.cell(row=4, column=1, value="Total Sessions:")
    ws_stats.cell(row=4, column=2, value=len(schedule))
    ws_stats.cell(row=5, column=1, value="Total Groups:")
    ws_stats.cell(row=5, column=2, value=len(groups_in_schedule))
    ws_stats.cell(row=6, column=1, value="Total Faculty:")
    ws_stats.cell(row=6, column=2, value=len(faculties_in_schedule))
    ws_stats.cell(row=7, column=1, value="Total Rooms Used:")
    ws_stats.cell(row=7, column=2, value=len(rooms_in_schedule))

    # Session Distribution by Group
    ws_stats.cell(row=9, column=1, value="👥 SESSIONS BY GROUP").font = Font(bold=True, size=14)
    ws_stats.cell(row=10, column=1, value="Group").font = Font(bold=True)
    ws_stats.cell(row=10, column=2, value="Sessions").font = Font(bold=True)
    row = 11
    for g in sorted(groups_in_schedule):
        count = sum(1 for info in schedule.values() if info["meta"]["group"] == g)
        ws_stats.cell(row=row, column=1, value=g)
        ws_stats.cell(row=row, column=2, value=count)
        row += 1

    # Session Distribution by Faculty
    ws_stats.cell(row=row+1, column=1, value="👨‍🏫 SESSIONS BY FACULTY").font = Font(bold=True, size=14)
    ws_stats.cell(row=row+2, column=1, value="Faculty").font = Font(bold=True)
    ws_stats.cell(row=row+2, column=2, value="Sessions").font = Font(bold=True)
    row += 3
    for f in sorted(faculties_in_schedule):
        count = sum(1 for info in schedule.values() if info["meta"]["faculty"] == f)
        ws_stats.cell(row=row, column=1, value=f)
        ws_stats.cell(row=row, column=2, value=count)
        row += 1

    # Adjust column widths for stats
    ws_stats.column_dimensions['A'].width = 25
    ws_stats.column_dimensions['B'].width = 15

    # ===== CLASH ANALYSIS =====
    if clashes:
        ws_clash = wb.create_sheet(title="Clash_Analysis")
        ws_clash.cell(row=1, column=1, value="CLASH ANALYSIS REPORT").font = Font(bold=True, size=16)
        ws_clash.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        # Check if cells are already merged before merging
        merge_range = "A1:C1"
        if not any(merge_range == str(mr) for mr in ws_clash.merged_cells.ranges):
            ws_clash.merge_cells(start_row=1, end_row=1, start_column=1, end_column=3)

        row = 3
        total_clashes = 0
        for clash_type, clash_list in clashes.items():
            if clash_list:
                ws_clash.cell(row=row, column=1, value=f"{clash_type.upper()} CLASHES:").font = Font(bold=True, color="FF0000")
                row += 1
                for item in clash_list:
                    ws_clash.cell(row=row, column=1, value=f"• {str(item)}")
                    row += 1
                    total_clashes += 1
            else:
                ws_clash.cell(row=row, column=1, value=f"✅ No {clash_type} clashes detected.").font = Font(color="008000")
                row += 1
            row += 1

        # Summary
        ws_clash.cell(row=row, column=1, value=f"TOTAL CLASHES FOUND: {total_clashes}").font = Font(bold=True, size=14)
        if total_clashes == 0:
            ws_clash.cell(row=row, column=1).font = Font(bold=True, size=14, color="008000")

        ws_clash.column_dimensions['A'].width = 60
    try:
        wb.save(filename)
        print(f"✅ Timetable exported to {filename}")
    except Exception as e:
        print(f"❌ Failed to export timetable: {e}")
        input("Press Enter to continue...")
        export_schedule_to_excel(schedule, days, slots_per_day, filename, time_labels, semester_name, clashes, groups, faculties, room_matrix, section, dept, room_no, ay, section_incharge)
