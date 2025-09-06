from collections import defaultdict
from openpyxl.styles import PatternFill

def detect_clashes(schedule, slots_per_day, rooms, group_sizes=None):
    clashes = {"faculty": [], "group": [], "room": [], "room_capacity": []}
    faculty_occ = defaultdict(list)
    group_occ = defaultdict(list)
    room_occ = defaultdict(list)
    session_type_count = defaultdict(int)
    room_map = {r['name']: r for r in rooms}
    lab_available = sum(1 for r in rooms if 'lab' in r['name'].lower())
    lecture_available = len(rooms) - lab_available
    elective_available = len(rooms)

    for sid, info in schedule.items():
        start, length, meta = info["start"], info["length"], info["meta"]
        room = info.get("room")
        name_lower = meta['name'].lower()
        sess_type = 'lab' if 'lab' in name_lower or 'project' in name_lower else 'elective' if 'elective' in name_lower else 'lecture'

        for offset in range(length):
            slot = start + offset
            faculty_occ[(meta['faculty'], slot)].append(sid)
            if "groups" in meta and meta["groups"]:
                for g in meta["groups"]:
                    group_occ[(g, slot)].append(sid)
            else:
                group_occ[(meta['group'], slot)].append(sid)
            if room:
                room_occ[(room, slot)].append(sid)
            session_type_count[(sess_type, slot)] += 1

        if room and group_sizes:
            g = meta['group']
            size = group_sizes.get(g, 0)
            cap = room_map.get(room, {}).get('capacity', 0)
            if size > cap:
                clashes['room_capacity'].append((sid, g, room, size, cap))

    for (faculty, slot), sids in faculty_occ.items():
        if len(sids) > 1: clashes["faculty"].append((faculty, slot, sids))
    for (group, slot), sids in group_occ.items():
        if len(sids) > 1: clashes["group"].append((group, slot, sids))
    for (room, slot), sids in room_occ.items():
        if len(sids) > 1: clashes["room"].append((room, slot, sids))

    for (sess_type, slot), count in session_type_count.items():
        if sess_type == 'lab' and count > lab_available: clashes['room_capacity'].append((sess_type, slot, count, lab_available))
        elif sess_type == 'lecture' and count > lecture_available: clashes['room_capacity'].append((sess_type, slot, count, lecture_available))
        elif sess_type == 'elective' and count > elective_available: clashes['room_capacity'].append((sess_type, slot, count, elective_available))

    return clashes

def highlight_clashes(ws, matrix, start_row=3, start_col=2, threshold=1):
    for i, row in enumerate(matrix):
        for j, val in enumerate(row):
            cell = ws.cell(row=start_row+i, column=start_col+j)
            if val > threshold:
                cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
